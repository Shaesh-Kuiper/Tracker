import sys
from pathlib import Path


def norm(s):
    return ''.join(ch for ch in str(s).strip().lower() if ch.isalnum())


def find_header_indexes(header_row, wanted_aliases):
    """
    header_row: list of header strings
    wanted_aliases: dict canonical_key -> set of normalized aliases
    returns dict canonical_key -> index
    """
    norm_header = [norm(h) for h in header_row]
    found = {}
    for key, aliases in wanted_aliases.items():
        idx = None
        for i, h in enumerate(norm_header):
            if h in aliases:
                idx = i
                break
        if idx is None:
            raise KeyError(f"Missing expected column for {key}. Headers: {header_row}")
        found[key] = idx
    return found


def load_leetcode_map(leetcode_xlsx: Path):
    import openpyxl
    wb = openpyxl.load_workbook(leetcode_xlsx, data_only=True)
    ws = wb.worksheets[0]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    aliases = {
        'reg': {"registernumber", "registrationnumber", "regno", "registerno"},
        'link': {"leetcodelink", "leetcode", "leetcodeprofile", "leetcodelinkurl"},
    }
    idx = find_header_indexes(header, aliases)
    reg_i, link_i = idx['reg'], idx['link']

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        reg = row[reg_i] if reg_i < len(row) else None
        link = row[link_i] if link_i < len(row) else None
        if reg is None or link in (None, ""):
            continue
        key = str(int(reg)) if isinstance(reg, (int, float)) else str(reg).strip()
        mapping[key] = str(link).strip()
    return mapping


def update_marquee(marquee_xlsx: Path, mapping: dict):
    import openpyxl
    wb = openpyxl.load_workbook(marquee_xlsx)
    # Prefer sheet named 'Students' if present
    ws = wb['Students'] if 'Students' in wb.sheetnames else wb.worksheets[0]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    aliases = {
        'reg': {"registernumber", "registrationnumber", "regno", "registerno"},
        'leetcode': {"leetcodelink", "leetcode", "leetcodeprofile", "leetcodelinkurl"},
        'name': {"name"},
    }
    idx = find_header_indexes(header, {
        'reg': aliases['reg'],
        'leetcode': aliases['leetcode'],
    })
    reg_i, lc_i = idx['reg'], idx['leetcode']

    updates = 0
    misses = 0
    for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
        reg_cell = row[reg_i]
        lc_cell = row[lc_i]
        reg_val = reg_cell.value
        key = None
        if reg_val is None:
            misses += 1
            continue
        if isinstance(reg_val, (int, float)):
            key = str(int(reg_val))
        else:
            key = str(reg_val).strip()

        link = mapping.get(key)
        if not link:
            misses += 1
            continue
        # If already same link, skip counting as update
        if str(lc_cell.value).strip() if lc_cell.value is not None else None:
            # Only update if different or empty
            existing = str(lc_cell.value).strip()
            if existing == link:
                continue
        lc_cell.value = link
        updates += 1

    wb.save(marquee_xlsx)
    return updates, misses


def main():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("This script requires openpyxl. Install with: python -m pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    base = Path('C:/Work/Tracker')
    leet = base / 'LeetCode.xlsx'
    marq = base / 'Marquee_Students.xlsx'
    if not leet.exists():
        print(f"Missing {leet}")
        sys.exit(1)
    if not marq.exists():
        print(f"Missing {marq}")
        sys.exit(1)

    # Backup marquee first
    backup = base / 'Marquee_Students.backup.xlsx'
    if not backup.exists():
        backup.write_bytes(marq.read_bytes())

    mapping = load_leetcode_map(leet)
    updates, misses = update_marquee(marq, mapping)
    print(f"Updated rows: {updates}")
    print(f"No-match/Skipped rows: {misses}")


if __name__ == '__main__':
    main()

