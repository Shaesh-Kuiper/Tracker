import sys
from pathlib import Path

def main():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not found. Please install with: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    base = Path('C:/Work/Tracker')
    files = ['LeetCode.xlsx', 'Marquee_Students.xlsx']
    for fname in files:
        f = base / fname
        if not f.exists():
            print(f"Missing file: {f}")
            continue
        print(f"=== {fname} ===")
        wb = openpyxl.load_workbook(f, data_only=True)
        for ws in wb.worksheets:
            # Read header row (first non-empty row)
            header = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() if c is not None else '' for c in row]
            # Trim trailing empties
            while header and header[-1] == '':
                header.pop()
            print(f"Sheet: {ws.title}")
            print(f"Columns ({len(header)}): {header}")
            # Print first 2 data rows for preview
            preview_rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=3, values_only=True), start=1):
                values = list(row[:len(header)]) if header else list(row)
                preview_rows.append(values)
            for idx, prow in enumerate(preview_rows, start=1):
                print(f"Row{idx}: {prow}")
        print()

if __name__ == '__main__':
    main()

