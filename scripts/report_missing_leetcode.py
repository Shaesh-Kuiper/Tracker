from pathlib import Path
import openpyxl

base = Path('C:/Work/Tracker')
marq = base / 'Marquee_Students.xlsx'
wb = openpyxl.load_workbook(marq, data_only=True)
ws = wb['Students'] if 'Students' in wb.sheetnames else wb.worksheets[0]
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
cols = {str(h).strip(): i for i, h in enumerate(header)}
name_i = cols.get('Name')
reg_i = cols.get('Registration Number')
lc_i = cols.get('LeetCode link')

missing = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[name_i] if name_i is not None and name_i < len(row) else ''
    reg = row[reg_i] if reg_i is not None and reg_i < len(row) else ''
    link = row[lc_i] if lc_i is not None and lc_i < len(row) else ''
    if link in (None, ''):
        missing.append((name, reg))

if missing:
    print('Rows with missing LeetCode link:')
    for name, reg in missing:
        print(f"- Name={name} Reg={reg}")
else:
    print('All rows have LeetCode links.')

