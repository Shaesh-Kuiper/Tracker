import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MARQUEE_PATH = r'C:\\Work\\Tracker\\Marquee.xlsx'
SAMPLE_PATH = r'C:\\Work\\Tracker\\sample_upload_excel.xlsx'


def find_header_row(ws, max_search_rows=20):
    max_r = min(ws.max_row or 1, max_search_rows)
    for r in range(1, max_r + 1):
        vals = [c.value for c in ws[r]]
        if any(v is not None and str(v).strip() != '' for v in vals):
            return r
    return 1


def extract_headers(ws):
    hr = find_header_row(ws)
    headers = []
    for c in ws[hr]:
        v = c.value
        if v is None:
            continue
        s = str(v).strip()
        if s == '':
            continue
        headers.append((c.col_idx, s, c))
    return hr, headers


def normalize(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(s).strip().lower())


def main():
    print('Loading workbooks...')
    wb_m = load_workbook(MARQUEE_PATH)
    wb_s = load_workbook(SAMPLE_PATH)

    ws_m = wb_m.worksheets[0]
    ws_s = wb_s.worksheets[0]

    print(f'Source sheet (Marquee): {ws_m.title}')
    print(f'Template sheet (Sample): {ws_s.title}')

    sample_header_row, sample_headers = extract_headers(ws_s)
    marquee_header_row, marquee_headers = extract_headers(ws_m)

    sample_header_names = [h[1] for h in sample_headers]
    marquee_header_names = [h[1] for h in marquee_headers]

    print('Sample headers:')
    print(' | '.join(sample_header_names))
    print('Marquee headers:')
    print(' | '.join(marquee_header_names))

    # Build mapping from normalized header -> column index for marquee
    marquee_map = {}
    for col_idx, name, cell in marquee_headers:
        key = normalize(name)
        if key and key not in marquee_map:
            marquee_map[key] = col_idx

    # Prepare new sheet name
    base_name = 'Upload_Format'
    new_name = base_name
    suffix = 1
    existing_names = {ws.title for ws in wb_m.worksheets}
    while new_name in existing_names:
        suffix += 1
        new_name = f'{base_name}_{suffix}'

    ws_new = wb_m.create_sheet(new_name)
    print(f'Created new sheet: {new_name}')

    # Write headers and copy styles/widths from template header
    for tgt_pos, (s_col_idx, header_text, sample_cell) in enumerate(sample_headers, start=1):
        cell = ws_new.cell(row=1, column=tgt_pos, value=header_text)
        if sample_cell is not None and sample_cell.has_style:
            cell.font = copy(sample_cell.font)
            cell.fill = copy(sample_cell.fill)
            cell.border = copy(sample_cell.border)
            cell.alignment = copy(sample_cell.alignment)
            cell.number_format = sample_cell.number_format
            cell.protection = copy(sample_cell.protection)
        # Copy column width
        col_letter_src = get_column_letter(s_col_idx)
        col_letter_tgt = get_column_letter(tgt_pos)
        width = ws_s.column_dimensions.get(col_letter_src).width if ws_s.column_dimensions.get(col_letter_src) else None
        if width is None:
            width = 15
        ws_new.column_dimensions[col_letter_tgt].width = width

    # Determine number formats for data columns based on sample (first non-empty below header)
    column_num_formats = {}
    for tgt_pos, (s_col_idx, header_text, sample_cell) in enumerate(sample_headers, start=1):
        fmt = None
        for r in range(sample_header_row + 1, min(ws_s.max_row, sample_header_row + 50) + 1):
            sc = ws_s.cell(row=r, column=s_col_idx)
            if sc.value is not None and str(sc.value) != '':
                if sc.has_style:
                    fmt = sc.number_format
                break
        if not fmt:
            fmt = 'General'
        column_num_formats[tgt_pos] = fmt

    # Build header mapping from sample headers -> marquee columns
    # Include basic synonym support for common field names
    synonyms = {
        'registrationnumber': ['registerno', 'registerno', 'regno', 'registrationno', 'regnumber', 'regdno', 'regid'],
        'geeksforgeekslink': ['gfglink', 'geeksforgeeks', 'gfgurl', 'geeksforgeeksurl', 'gfgprofile', 'gfg'],
        'leet codelink': ['leetcode', 'leetcodeurl', 'leetcodeprofile', 'leetcodeid', 'leetcodelink', 'leet_code_link'],
        'codecheflink': ['codechef', 'codechefurl', 'codechefprofile', 'codecheflink'],
        'name': ['studentname', 'fullname', 'full_name', 'candidate', 'applicant', 'name']
    }
    # fix keys: normalize synonym dict keys
    synonyms = {normalize(k): [normalize(v) for v in vals] for k, vals in synonyms.items()}

    mapping = []
    for tgt_pos, (_, header_text, _) in enumerate(sample_headers, start=1):
        key = normalize(header_text)
        src_col = marquee_map.get(key)
        if not src_col and key in synonyms:
            for alt in synonyms[key]:
                if alt in marquee_map:
                    src_col = marquee_map[alt]
                    break
        mapping.append((tgt_pos, header_text, src_col))

    print('Header mapping (Sample -> Marquee):')
    for tgt_pos, header_text, src_col in mapping:
        print(f'  [{tgt_pos}] {header_text}  <=  ' + (f'col {src_col}' if src_col else '(no match)'))

    # Copy data rows from marquee into new sheet according to mapping
    start_row_src = marquee_header_row + 1
    start_row_tgt = 2
    max_src_row = ws_m.max_row
    rows_written = 0
    for r in range(start_row_src, max_src_row + 1):
        # Detect empty row in source
        row_vals = [ws_m.cell(row=r, column=col_idx).value for col_idx in range(1, ws_m.max_column + 1)]
        if all(v is None or str(v).strip() == '' for v in row_vals):
            continue
        tgt_r = start_row_tgt + rows_written
        for tgt_pos, header_text, src_col in mapping:
            c = ws_new.cell(row=tgt_r, column=tgt_pos)
            if src_col:
                v = ws_m.cell(row=r, column=src_col).value
                c.value = v
            c.number_format = column_num_formats.get(tgt_pos, 'General')
        rows_written += 1

    # Freeze header and add autofilter
    ws_new.freeze_panes = 'A2'
    if sample_headers:
        end_col_letter = get_column_letter(len(sample_headers))
        ws_new.auto_filter.ref = f'A1:{end_col_letter}1'

    wb_m.save(MARQUEE_PATH)
    print(f'Completed. Wrote {rows_written} data row(s) into sheet "{new_name}" of {MARQUEE_PATH}.')


if __name__ == '__main__':
    main()
